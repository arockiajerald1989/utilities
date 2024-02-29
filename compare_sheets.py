import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# Specify filenames and sheet names
filename1 = 'FTP.xlsx'  # Replace with the actual filename for V1
sheet_name1 = 'Sheet1'  # Replace with the actual sheet name for V1 if different
filename2 = 'TS.xlsx'  # Replace with the actual filename for V2
sheet_name2 = 'Sheet1'  # Replace with the actual sheet name for V2 if different

# Read the Excel files, stopping at the first empty row in each
df1 = pd.read_excel(filename1, sheet_name=sheet_name1, header=0, nrows=None)
df2 = pd.read_excel(filename2, sheet_name=sheet_name2, header=0, nrows=None)

# Total the Hours by FullName and ID for each file separately, ensuring MultiIndex
total_hours_ver1 = df1.groupby(['FullName', 'ID'])['Total Hours'].sum().reset_index()  # Reset index to create MultiIndex
total_hours_ver2 = df2.groupby(['FullName', 'ID'])['Total Hours'].sum().reset_index()

# Merge the total_hours DataFrames
df_merged = pd.merge(total_hours_ver1, total_hours_ver2, on=['ID'], suffixes=('_FTP', '_TS'), how='outer')

# Add the Difference Hours column
df_merged['Difference'] = df_merged['Total Hours_FTP'] - df_merged['Total Hours_TS']

# Rearrange columns in the desired order
df_merged = df_merged[['ID', 'FullName_TS', 'Total Hours_FTP', 'Total Hours_TS', 'Difference']]
print(df_merged)

# Write the merged DataFrame to a new Excel sheet
df_merged.to_excel('TEST.xlsx', index=False)

# Create a new Excel workbook using openpyxl
workbook = Workbook()
worksheet = workbook.active

# Write column headers to the worksheet
worksheet.append(df_merged.columns.tolist())

# Iterate through rows and apply color formatting
for row_index, row in df_merged.iterrows():
    worksheet.append(row.tolist())
    if row['Difference'] != 0:  # Apply red color to non-zero cells
        cell = worksheet.cell(row=row_index + 2, column=6)  # Adjust column index if needed
        cell.font = Font(color='FF0000')  # Red font color
        cell.fill = PatternFill(bgColor='FFC7CE')  # Light red background color

# Save the workbook
workbook.save('TEST.xlsx')
